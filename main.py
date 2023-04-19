
# %appdata%  Roaming\gspread\service_account.json
# запрос на указание файла с данными
import time
from tkinter.filedialog import askopenfilename
from datetime import datetime
import openpyxl
import gspread  # pip install openpyxl gspread


import pip
pip.main(["install", "gspread"])
pip.main(["install", "openpyxl"])


# from imp import find_module

# def checkPythonmod(mod):
#     try:
#         op = find_module(mod)
#         return True
#     except ImportError:
#         return False
# try:
#   import openpyxl
# except ImportError as e:
#   pip.main(["install", "openpyxl"])
#   import openpyxl

# try:
#   import gspread
# except ImportError as e:
#   pip.main(["install", "gspread"])
#   import gspread

# if python -c 'import pkgutil; exit(not pkgutil.find_loader("pandas"))'; then
#     echo 'pandas found'
# else
#     echo 'pandas not found'
# fi

# starting time
start = time.time()


# URL = "orders_monitoring_2023_04_18_13_07_20364910.xlsx"

# запрос на указание файла с данными
URL = askopenfilename()

# FILE_NAME = 'test.xlsx'
# wb = openpyxl.reader.excel.load_workbook(filename=FILE_NAME,data_only=True)

# открываем локальную таблицу менять data_only если надо формулы
wb = openpyxl.reader.excel.load_workbook(filename=URL, data_only=True)
# print(wb.sheetnames) # показать имя листа
wb.active = 0  # активировать самый левый лист в книге
# сохранить в переменную для дальнейшей работы с ним
sheetLocal = wb.active


# функция чтобы найти число в конце строки
def atlestCharToInt(cellValue):
    # если применяем к пустой строке,
    """Function printing python version."""
    if (cellValue == '') or (cellValue == 0) or (cellValue is None):
        SS = 0
        return SS  # то вернуть пустую строку
    else:
        # lengthLine = len(cellValue.velue)
        # забираем 3 последних символа, оберзаем пробелы
        SS = cellValue[len(cellValue)-3:len(cellValue)].strip()
        # полученную строку переводим в числовой тип данных
        return int(SS)

# a = "Питающихся: 25Комплекты:- Завтрак 1-4 класс: 25"
# # СТрока для тестов нахождения последних символов
# s = atlestCharToInt(a); # результат функции для нахождения последних цифры в строке


def split(arr, size):  # функция для разбиения листа на под листы, чтбоы отправить на страницу
    """Function printing python version."""
    arrs = []
    while len(arr) > size:
        pice = arr[:size]
        arrs.append(pice)
        arr = arr[size:]
    arrs.append(arr)
    return arrs


# Строка для инициализации, что начали работать с файлами
print('Start work')

# авторизация гугл ака, по json файлу
gc = gspread.service_account()
# открытие гугл таблицы с именем "PythonSheets"
sh = gc.open("Питание Лицей")
# Выбрать ппервый активный у нее лист
GoogleSheets = sh.sheet1

array = []
for i in range(50, 54, 1):  # формирование обед и полдник у 2АБВГ
    array.append(atlestCharToInt(sheetLocal['D' + str(i)].value))
    array.append(atlestCharToInt(sheetLocal['E' + str(i)].value))
# пакуем в подлисты по 2 для вставки в гугл таблицу
x = split(array, 2)

array = []
for i in range(5, 18, 1):  # завтраки 1АБВГД, 2АБВГ 3АБВГ класс
    array.append(atlestCharToInt(sheetLocal['C' + str(i)].value))
y = split(array, 1)

array = []
for i in range(18, 22, 1):  # обеды 4АБВГ класс
    array.append(atlestCharToInt(sheetLocal['D' + str(i)].value))
z = split(array, 1)

array = []
for i in range(22, 26, 1):  # завтраки 5АБВГ класс
    array.append(atlestCharToInt(sheetLocal['C' + str(i)].value))
v = split(array, 1)

array = []
for i in range(26, 33, 1):  # обеды 6АБВГ 7АБВ класс
    array.append(atlestCharToInt(sheetLocal['D' + str(i)].value))
b = split(array, 1)

array = []
for i in range(33, 36, 1):  # завтраки 8АБВ класс
    array.append(atlestCharToInt(sheetLocal['C' + str(i)].value))
n = split(array, 1)

array = []
for i in range(36, 45, 1):  # завтраки 9АБВ 10АБВ 11АБВ класс
    array.append(atlestCharToInt(sheetLocal['C' + str(i)].value))
nn9 = split(array, 1)


gg = []  # формирование обед и полдник у 1АБВГД
classLastNames = ['А', 'Б', 'В', 'Г', 'Д']

for i in classLastNames:  # созданеи нужной последовательности ГПД в 1-х классах
    for j in range(45, 50, 1):
        s = sheetLocal['B' + str(j)].value
        if s[len(s)-1] == i:
            XY = atlestCharToInt(sheetLocal['D' + str(j)].value)
            gg.append(XY)
            XY = atlestCharToInt(sheetLocal['E' + str(j)].value)
            gg.append(XY)
            break

gg = split(gg, 2)
# print(f"gg = {gg}")

currentDay = datetime.now().day
if currentDay <= 10:
    currentDay = f'0{currentDay}'
currentMonth = datetime.now().month
if currentMonth <= 10:
    currentMonth = f'0{currentMonth}'

currentYear = datetime.now().year
today = f'{currentDay}.{currentMonth}.{currentYear}'
#
# today = "=TODAY()"

# print('today ' + str(today))
# today = split(today,1)
# Ввод информации в гугл таблицу

GoogleSheets.update("E1", today)
GoogleSheets.batch_update([{
    'range': 'C8:D11',  # диапазон куда грузим
    'values': x,  # загружаем обед и полдник у 2АБВГ
}, {
    'range': 'C3:D7',  # диапазон куда грузим
    'values': gg,  # загружаем обед и полдник у 1АБВГД
}, {
    'range': 'G35:G43',  # диапазон куда грузим
    'values': nn9,  # загружаем обед и полдник у 1АБВГД
}, {
    'range': 'L3:L15',  # диапазон куда грузим
    'values': y,  # загружаем завтраки 1АБВГД, 2АБВГ 3АБВГ класс
}, {
    'range': 'M16:M19',  # диапазон куда грузим
    'values': z,  # загружаем обеды 4АБВГ класс
}, {
    'range': 'G21:G24',  # диапазон куда грузим
    'values': v,  # загружаем завтраки 5АБВГ класс
}, {
    'range': 'I25:I31',  # диапазон куда грузим
    'values': b,  # загружаем обеды 6АБВГ 7АБВ класс
}, {
    'range': 'G32:G34',  # диапазон куда грузим
    'values': n,  # загружаем завтраки 8АБВ класс
}])

# end time
end = time.time()

print(f"Execution time of the program is- {end-start:5.3f} s.")
